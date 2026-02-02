import streamlit as st
import os
import json
import pickle
import datetime
import pandas as pd
import numpy as np  # Added for QC logic
import faiss
import io
import time
from typing import TypedDict, List, Annotated
from sentence_transformers import SentenceTransformer
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ---------------- LangGraph ----------------
from langchain_core.messages import AnyMessage
from langgraph.graph import StateGraph, START, END
from langgraph.graph.message import add_messages

# ---------------- LLM ----------------
from openai import OpenAI

# ==========================================================
# 1. PAGE CONFIG & STYLES
# ==========================================================
st.set_page_config(page_title="Defense Agentic Extraction", page_icon="🛡️", layout="wide")

st.markdown("""
<style>
    .agent-box {
        padding: 10px;
        border-radius: 8px;
        text-align: center;
        margin: 5px;
        font-weight: bold;
        border: 1px solid #ddd;
    }
    .agent-waiting { background-color: #f0f2f6; color: #888; }
    .agent-active { background-color: #d1e7dd; color: #0f5132; border: 2px solid #198754; box-shadow: 0 0 10px rgba(25,135,84,0.3); }
    .agent-done { background-color: #cfe2ff; color: #084298; }
</style>
""", unsafe_allow_html=True)

# ==========================================================
# 2. PROMPTS & CONFIG
# ==========================================================

REQUIRED_COLUMNS = [
    "Customer Region", "Customer Country", "Customer Operator", 
    "Supplier Region", "Supplier Country", "Domestic Content", 
    "Market Segment", "System Type (General)", "System Type (Specific)", 
    "System Name (General)", "System Name (Specific)", "System Piloting", 
    "Supplier Name", "Program Type", "Expected MRO Contract Duration (Months)", 
    "Quantity", "Value Certainty", "Value (Million)", "Currency", 
    "Value (USD$ Million)", "Value Note\n(If Any)", "G2G/B2G", 
    "Signing Month", "Signing Year", "Description of Contract", 
    "Additional Notes (Internal Only)", "Source Link(s)", 
    "Contract Date", "Reported Date (By SGA)"
]

GEOGRAPHY_PROMPT = """
You are a Defense Geography Analyst. 
Extract the Customer Country, Customer Operator, and Supplier Country from the text.

STRICT RULES:
1. **Customer Country**: 
   - Identify the government/nation PAYING for or RECEIVING the goods.
   - If "Foreign Military Sales (FMS)" is mentioned, look for the specific country name (e.g., "FMS to Japan").
   - Do NOT assume the "Work Location" is the Customer. (e.g., Work in Alabama for a contract supporting the UK -> Customer is UK).

2. **Customer Operator**:
   - Extract the specific service branch (e.g., "Navy", "Air Force", "Army", "Coast Guard", "Marines").
   - If a specific foreign military branch is named (e.g., "Royal Australian Air Force"), extract that.

3. **Supplier Country**:
   - Identify the country where the Supplier Company's headquarters is located.

Return JSON ONLY:
{
  "Customer Country": "...",
  "Customer Operator": "...",
  "Supplier Country": "..."
}
"""

SYSTEM_CLASSIFIER_PROMPT = """
You are a Senior Defense System Classification Analyst.

1. **REFERENCE TAXONOMY**:
{taxonomy_reference}

2. **RULE BOOK OVERRIDES**:
{rule_book_overrides}

3. **TASK**:
   - Classify the system described in the contract into **Market Segment**, **System Type (General)**, and **System Name**.
   - **CRITICAL**: If "ITEM_FOCUS" is provided, classify THAT specific item. If empty, classify the main system in the text.
   - Use the "RAG Examples" provided to guide your choice if the text is similar.

4. **CLASSIFICATION RULES**:
   - **Generic IT/Enterprise Software**: If the contract is for generic office software (e.g., Microsoft 365, DoD ESI), cloud services, or non-tactical IT, classify Market Segment as **"Unknown"** or **"Not Applicable"**.
   - **Air vs Navy**: If the system is an Aircraft (e.g., P-8, E-2D, F-35), Market Segment is **"Air Platforms"**, even if the customer is the Navy.
   - **Ship/Submarine**: Market Segment is **"Naval Platforms"**.

5. **SYSTEM NAME EXTRACTION**:
   - **System Name (General)**: The **Host Platform** or **Class** (e.g., "E-2D Advanced Hawkeye", "Arleigh Burke-class", "Los Angeles-class").
   - **System Name (Specific)**: The **Specific Subject** of the contract.
     - If it's a specific ship/aircraft instance: Extract the name/hull number (e.g., "USS Pinckney (DDG-91)", "USS Hartford (SSN-768)", "USNS Robert Ballard (T-AGS 67)").
     - If it's a service/mod description: Extract the description (e.g., "Extend Services and Adds Hours...", "Depot Modernization Period").
     - If it's a component: Extract the component name.

6. **OUTPUT RULES**:
   - Return ONLY a FLAT JSON object.
   - Evidence must be copied EXACTLY from the text.
   - If evidence is not present, output "Not Found".

Return JSON:
{
  "Market Segment": "...",
  "Market Segment Evidence": "...",
  "Market Segment Reason": "...",
  
  "System Type (General)": "...",
  "System Type (General) Evidence": "...",
  "System Type (General) Reason": "...",

  "System Type (Specific)": "...",
  "System Type (Specific) Evidence": "...",
  "System Type (Specific) Reason": "...",

  "System Name (General)": "...",
  "System Name (General) Evidence": "...",
  "System Name (General) Reason": "...",

  "System Name (Specific)": "...",
  "System Name (Specific) Evidence": "...",
  "System Name (Specific) Reason": "...",

  "Confidence": "High/Medium/Low"
}
"""

CONTRACT_EXTRACTOR_PROMPT = """
You are a Defense Contract Financial Analyst.

1. **TASK**: Extract supplier, program type, financial certainty, FMS status, completion date, currency, and SIGNING DATE details.
2. **PROGRAM TYPE ENUM**:
    {program_type_enum}

3. **STRICT RULES**:
    - **Supplier Name**: Extract the **Clean Entity Name**. Include the **Major Division** if specified (e.g., "General Dynamics Electric Boat", "Northrop Grumman Aerospace"). Do not include legal suffixes like "Corp", "Inc", "L.P." unless part of the brand.
    - **Program Type**:
      - **MRO/Support**: Includes "depot modernization", "maintenance", "overhaul", "repair", "sustainment", "logistics support".
      - **Procurement**: Includes "production", "manufacture", "delivery" of new hardware.
      - **RDT&E**: Research, development, prototyping.
    - **Value Certainty**: 
      - "Confirmed" for definite contracts/mods.
      - "Estimated" for IDIQ ceilings, "potential value", or "maximum value".
    - **G2G/B2G**: "G2G" ONLY if "Foreign Military Sales" (FMS) is explicitly mentioned. Otherwise "B2G".
    - **Value Note**: Capture notes about IDIQs, options, or ceilings.

4. **DATE EXTRACTION RULES (CRITICAL)**:
    - **Signing Month**: The month the contract was **SIGNED**. 
      - NOT the date of Letter of Intent (LoI), first payment, or delivery.
      - If the text says "announced today", use the "Reference Date" provided in the prompt to determine the month.
    - **Signing Year**: The year the contract was **SIGNED**.
    - Output full month name (e.g., "January") and 4-digit year (e.g., "2024").

5. **FINANCIAL EXTRACTION RULES (SOP)**:
    - **Primary Amount**: Look for the FIRST contract value mentioned in the text.
    - **Value (Million)**: Extract this value as a number in MILLIONS. 
      - If the text says "$500,000", output 0.500.
      - If the text says "$1.2 Billion", output 1200.000.
      - If the text says "$45 Million", output 45.000.
    - **Currency**: Extract the ISO 3-letter currency code (e.g., USD, EUR, GBP).

Return JSON ONLY:
{
  "program_type": "...",
  "currency_code": "...",
  "value_in_millions": 0.000,
  "value_certainty": "...",
  "completion_date_text": "...",
  "g2g_b2g": "...",
  "value_note": "...",
  "extracted_supplier": "...",
  "signing_month": "...",
  "signing_year": "..."
}
"""

# ==========================================================
# 3. CACHED RESOURCES (KB & LLM)
# ==========================================================

@st.cache_resource
def load_kb_resources(kb_dir):
    try:
        if not os.path.exists(os.path.join(kb_dir, "system_kb.faiss")):
            return None
        
        index = faiss.read_index(os.path.join(kb_dir, "system_kb.faiss"))
        with open(os.path.join(kb_dir, "system_kb_meta.pkl"), "rb") as f:
            meta = pickle.load(f)
        embedder = SentenceTransformer("sentence-transformers/all-mpnet-base-v2")
        return {"index": index, "meta": meta, "embedder": embedder}
    except Exception as e:
        return None

def get_kb_hit(resources, text):
    if not resources: return {}, 0.0, None
    emb = resources["embedder"].encode([text], normalize_embeddings=True).astype("float32")
    scores, idxs = resources["index"].search(emb, 1)
    if idxs[0][0] < 0: return {}, 0.0, None
    return resources["meta"][idxs[0][0]], float(scores[0][0]), idxs[0][0]

def call_llm_api(prompt, api_key, provider="foundry", max_tokens=500):
    try:
        if provider == "openrouter":
            client = OpenAI(api_key=api_key, base_url="https://openrouter.ai/api/v1")
            model = "openai/gpt-4o-mini"
        else:
            client = OpenAI(api_key=f'{api_key}:agentic', base_url="https://llmfoundry.straive.com/openai/v1/")
            model = "gpt-4o-mini"

        response = client.chat.completions.create(
            model=model,
            messages=[{"role": "user", "content": prompt}],
            temperature=0,
            max_tokens=max_tokens,
            response_format={"type": "json_object"},
        )
        return json.loads(response.choices[0].message.content)
    except Exception:
        return {}

# ==========================================================
# 4. AGENT LOGIC
# ==========================================================

class AgentState(TypedDict):
    text: str; date: str; url: str
    kb_meta: dict; kb_score: float; kb_mode: str; kb_row_id: int | None
    row: dict; rows: list
    messages: Annotated[List[AnyMessage], add_messages]

def build_workflow(kb_resources, foundry_key, openrouter_key, rag_thresholds):
    
    def mode_from_score(score):
        if score >= rag_thresholds["strong"]: return "KB_ONLY"
        if score >= rag_thresholds["medium"]: return "KB_GUIDED"
        return "LLM_ONLY"

    def kb_router(state: AgentState):
        meta, score, row_id = get_kb_hit(kb_resources, state["text"])
        return {"kb_meta": meta, "kb_score": score, "kb_mode": mode_from_score(score), "kb_row_id": row_id}

    def sourcing(state: AgentState):
        return {"row": {
            "Description of Contract": state["text"], "Contract Date": state["date"],
            "Source Link(s)": state["url"], "Reported Date (By SGA)": datetime.date.today().isoformat()
        }}

    def geography(state: AgentState):
        row = state["row"].copy()
        if state["kb_mode"] != "LLM_ONLY":
            for k in ["Customer Country", "Customer Region", "Customer Operator", "Supplier Country", "Supplier Region", "Domestic Content"]:
                if state["kb_meta"].get(k): row[k] = state["kb_meta"][k]
        else:
            api_key = foundry_key if foundry_key else openrouter_key
            provider = "foundry" if foundry_key else "openrouter"
            res = call_llm_api(GEOGRAPHY_PROMPT + "\n" + state["text"], api_key, provider)
            if res: row.update(res)
        return {"row": row}

    def system(state: AgentState):
        row = state["row"].copy()
        if state["kb_mode"] != "LLM_ONLY":
            for k in ["Market Segment", "System Type (General)", "System Type (Specific)", "System Name (General)", "System Name (Specific)", "System Piloting"]:
                if state["kb_meta"].get(k): row[k] = state["kb_meta"][k]
        else:
            api_key = foundry_key if foundry_key else openrouter_key
            provider = "foundry" if foundry_key else "openrouter"
            res = call_llm_api(SYSTEM_CLASSIFIER_PROMPT + "\n" + state["text"], api_key, provider)
            if res: row.update(res)
        return {"row": row}

    def contract(state: AgentState):
        row = state["row"].copy()
        context = f"Reference Date: {state['date']}\n\nContract Text: {state['text']}"
                
        if state["kb_mode"] != "LLM_ONLY":
            # KB Extraction Logic
            for k in ["Supplier Name", "Program Type", "Value (Million)", "Value (USD$ Million)", "Currency", 
                            "Value Certainty", "Quantity", "G2G/B2G", "Signing Month", "Signing Year", 
                            "Expected MRO Contract Duration (Months)"]:
                if state["kb_meta"].get(k): row[k] = state["kb_meta"][k]
                else:
                    # LLM Extraction Logic
                    api_key = foundry_key if foundry_key else openrouter_key
                    provider = "foundry" if foundry_key else "openrouter"
                    
                    res = call_llm_api(CONTRACT_EXTRACTOR_PROMPT + "\n" + context, api_key, provider)
                    
                    # --- NEW: SOP Financial Mapping ---
                    val_millions = float(res.get("value_in_millions", 0))
                    currency = res.get("currency_code", "USD").upper()
                    
                    # 1. Capture the direct value (Millions, 3 decimal places)
                    row["Value (Million)"] = round(val_millions, 3)
                    row["Currency"] = currency
                    
                    # 2. Handle USD Conversion for "Value (USD$ Million)"
                    if currency == "USD":
                        row["Value (USD$ Million)"] = round(val_millions, 3)
                    else:
                        # Basic conversion logic - you can expand this dictionary as needed
                        # Or use a live conversion library
                        rates = {"EUR": 1.08, "GBP": 1.27, "AUD": 0.65, "CAD": 0.74}
                        rate = rates.get(currency, 1.0) 
                        row["Value (USD$ Million)"] = round(val_millions * rate, 3)

                    # --- Keep existing fields ---
                    row["Supplier Name"] = res.get("extracted_supplier")
                    row["Program Type"] = res.get("program_type")
                    row["Value Certainty"] = res.get("value_certainty")
                    row["Value Note\n(If Any)"] = res.get("value_note")
                    row["G2G/B2G"] = res.get("g2g_b2g")
                    row["Signing Month"] = res.get("signing_month")
                    row["Signing Year"] = res.get("signing_year")
                    row["Expected MRO Contract Duration (Months)"] = res.get("mro_duration_months", "Not Applicable")
                    
                return {"rows": [row]}

    def evaluation(state: AgentState):
        evaluated = []
        for r in state["rows"]:
            row = r.copy()
            row["Accuracy Score"] = 90 if state["kb_mode"] == "KB_ONLY" else 70 if state["kb_mode"] == "KB_GUIDED" else 50
            row["Extraction Source"] = state["kb_mode"]
            evaluated.append(row)
        return {"rows": evaluated}

    def formatter(state: AgentState): return {}

    graph = StateGraph(AgentState)
    graph.add_node("KBRouter", kb_router)
    graph.add_node("Sourcing", sourcing)
    graph.add_node("Geography", geography)
    graph.add_node("System", system)
    graph.add_node("Contract", contract)
    graph.add_node("Evaluation", evaluation)
    graph.add_node("ExcelFormatter", formatter)

    graph.add_edge(START, "KBRouter")
    graph.add_edge("KBRouter", "Sourcing")
    graph.add_edge("Sourcing", "Geography")
    graph.add_edge("Geography", "System")
    graph.add_edge("System", "Contract")
    graph.add_edge("Contract", "Evaluation")
    graph.add_edge("Evaluation", "ExcelFormatter")
    graph.add_edge("ExcelFormatter", END)

    return graph.compile()

# ==========================================================
# 5. QC LOGIC (Added)
# ==========================================================
def run_qc_check(structured_file, raw_file):
    results = []
    
    try:
        # Load Data
        if structured_file.name.endswith('.csv'):
            structured_df = pd.read_csv(structured_file)
        else:
            structured_df = pd.read_excel(structured_file)
            
        if raw_file.name.endswith('.csv'):
            raw_df = pd.read_csv(raw_file)
        else:
            raw_df = pd.read_excel(raw_file)
            
    except Exception as e:
        return None, f"Error loading files: {str(e)}"
    
    # Get columns
    structured_cols = structured_df.columns.tolist()
    
    # Iterate
    for col in structured_cols:
        if col in raw_df.columns:
            s_vals = structured_df[col]
            r_vals = raw_df[col]
            
            match_count = 0
            mismatches = []
            
            rows_to_check = min(len(structured_df), len(raw_df))
            
            for i in range(rows_to_check):
                s_val = s_vals.iloc[i]
                r_val = r_vals.iloc[i]
                
                is_match = False
                
                # Check for NaNs
                if pd.isna(s_val) and pd.isna(r_val):
                    is_match = True
                elif pd.isna(s_val) or pd.isna(r_val):
                    is_match = False
                else:
                    # String compare
                    if isinstance(s_val, str) and isinstance(r_val, str):
                        is_match = str(s_val).strip() == str(r_val).strip()
                    # Numeric compare
                    elif isinstance(s_val, (int, float)) and isinstance(r_val, (int, float)):
                        is_match = np.isclose(s_val, r_val, equal_nan=True)
                    else:
                        is_match = s_val == r_val
                
                if is_match:
                    match_count += 1
                else:
                    mismatches.append(f"Row {i+2}: '{s_val}' vs '{r_val}'")
            
            accuracy = (match_count / rows_to_check) * 100 if rows_to_check > 0 else 0
            
            if mismatches:
                desc = "; ".join(mismatches[:3])
                if len(mismatches) > 3:
                    desc += f"; ... and {len(mismatches) - 3} more."
            else:
                desc = "No discrepancies"
            
            results.append({
                "Column Name": col,
                "Accuracy (%)": round(accuracy, 2),
                "Incorrect Description": desc
            })
        else:
            results.append({
                "Column Name": col,
                "Accuracy (%)": 0.0,
                "Incorrect Description": "Column not found in Raw Data"
            })
            
    return pd.DataFrame(results), None

# ==========================================================
# 6. UI LAYOUT & EXECUTION
# ==========================================================

st.title("🛡️ Defense Agentic Pipeline")

# Sidebar Navigation
st.sidebar.title("Navigation")
app_mode = st.sidebar.radio("Go to:", ["Extraction Pipeline", "Quality Control"])

# ----------------- EXTRACTION PIPELINE -----------------
if app_mode == "Extraction Pipeline":
    st.markdown("Extract defense contract data using a 7-Agent System with Real-Time Visualization.")
    
    with st.sidebar:
        st.header("⚙️ Settings")
        st.subheader("1. API Keys")
        foundry_key = st.text_input("LLM Foundry Token", type="password")
        openrouter_key = st.text_input("OpenRouter Key", type="password")
        
        st.subheader("2. Knowledge Base")
        kb_path = st.text_input("KB Folder Path", value=r"C:\Users\mukeshkr\Agentic-AI-Defense-Data-Extraction\notebook\system_kb_store")
        
        st.subheader("3. RAG Sensitivity")
        rag_strong = st.slider("Strong Match", 0.0, 1.0, 0.78)
        rag_medium = st.slider("Medium Match", 0.0, 1.0, 0.70)

    uploaded_file = st.file_uploader("📂 Upload Source Excel", type=["xlsx"])

    # UI Containers
    status_container = st.container()
    visualization_container = st.container()
    result_container = st.container()

    if uploaded_file and st.button("🚀 Start Extraction"):
        if not foundry_key and not openrouter_key:
            st.error("❌ Please enter at least one API Key.")
            st.stop()

        kb_resources = load_kb_resources(kb_path)
        app = build_workflow(kb_resources, foundry_key, openrouter_key, {"strong": rag_strong, "medium": rag_medium})
        
        df = pd.read_excel(uploaded_file)
        all_extracted_rows = []

        progress_bar = status_container.progress(0)
        current_status = status_container.empty()
        
        with visualization_container:
            st.write("#### 🤖 Agent Activity Monitor")
            cols = st.columns(7)
            agent_names = ["KBRouter", "Sourcing", "Geography", "System", "Contract", "Evaluation", "Formatter"]
            placeholders = [col.empty() for col in cols]

        def update_agents(active_idx):
            for i, name in enumerate(agent_names):
                if i < active_idx:
                    placeholders[i].markdown(f'<div class="agent-box agent-done">✅ {name}</div>', unsafe_allow_html=True)
                elif i == active_idx:
                    placeholders[i].markdown(f'<div class="agent-box agent-active">⚙️ {name}</div>', unsafe_allow_html=True)
                else:
                    placeholders[i].markdown(f'<div class="agent-box agent-waiting">{name}</div>', unsafe_allow_html=True)

        # Processing Loop
        for i, r in df.iterrows():
            progress_bar.progress((i + 1) / len(df))
            current_status.markdown(f"**Processing Row {i+1}/{len(df)}**: _{str(r.get('Contract Description', ''))[:50]}..._")
            
            state = {
                "text": str(r.get("Contract Description", "")),
                "date": str(r.get("Contract Date", "")),
                "url": str(r.get("Source URL", "")),
                "row": {}, "rows": [], "messages": []
            }

            step_mapping = {"KBRouter": 0, "Sourcing": 1, "Geography": 2, "System": 3, "Contract": 4, "Evaluation": 5, "ExcelFormatter": 6}
            
            # Temp storage for the row being processed
            processed_row = None

            try:
                for event in app.stream(state):
                    for node_name, output in event.items():
                        if node_name in step_mapping:
                            update_agents(step_mapping[node_name])
                            time.sleep(0.05) 
                        
                        # Capture rows explicitly from Evaluation node
                        if node_name == "Evaluation" and "rows" in output:
                            processed_row = output["rows"]

                update_agents(7)

                if processed_row:
                    all_extracted_rows.extend(processed_row)
                
                # Live Update in Result Container
                with result_container:
                    if all_extracted_rows:
                        live_df = pd.DataFrame(all_extracted_rows)
                        st.write(f"### Live Results ({len(all_extracted_rows)} Records)")
                        st.dataframe(live_df.tail(3), use_container_width=True)

            except Exception as e:
                st.error(f"Error processing row {i}: {e}")

        current_status.success("✅ Extraction Complete!")

        # Final Output
        if all_extracted_rows:
            df_raw = pd.DataFrame(all_extracted_rows)
            all_cols = set().union(*(d.keys() for d in all_extracted_rows))
            df_raw = df_raw.reindex(columns=sorted(all_cols), fill_value="")
            df_structured = df_raw.reindex(columns=REQUIRED_COLUMNS, fill_value="")

            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                df_structured.to_excel(writer, sheet_name="Structured Output", index=False)
                df_raw.to_excel(writer, sheet_name="Raw Data", index=False)
            
            buffer.seek(0)
            wb = load_workbook(buffer)
            if "Raw Data" in wb.sheetnames:
                ws = wb["Raw Data"]
                headers = [c.value for c in ws[1]]
                if "Accuracy Score" in headers:
                    idx = headers.index("Accuracy Score") + 1
                    green = PatternFill("solid", fgColor="C6EFCE")
                    yellow = PatternFill("solid", fgColor="FFEB9C")
                    red = PatternFill("solid", fgColor="F4CCCC")
                    for r_idx in range(2, ws.max_row + 1):
                        cell = ws.cell(row=r_idx, column=idx)
                        try:
                            val = int(cell.value)
                            cell.fill = green if val >= 85 else yellow if val >= 65 else red
                        except: pass
            
            final_buffer = io.BytesIO()
            wb.save(final_buffer)
            final_buffer.seek(0)

            st.markdown("---")
            st.subheader("📥 Download Final Results")
            st.download_button(
                label="Download Excel File",
                data=final_buffer,
                file_name="Processed_Defense_Data.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.warning("No data was extracted. Please check your API keys or input file format.")

# ----------------- QUALITY CONTROL -----------------
elif app_mode == "Quality Control":
    st.header("📊 Quality Control & Analytics")
    st.markdown("Compare the **Structured Output** against a **Verified/Raw Data** source to calculate accuracy and identify discrepancies.")
    
    qc_col1, qc_col2 = st.columns(2)
    with qc_col1:
        st.subheader("1. Upload Generated File")
        f_struct = st.file_uploader("Upload 'Structured Output' (CSV/Excel)", type=["csv", "xlsx"], key="u1")
        
    with qc_col2:
        st.subheader("2. Upload Verified File")
        f_raw = st.file_uploader("Upload 'Raw Data' / Ground Truth (CSV/Excel)", type=["csv", "xlsx"], key="u2")
        
    if f_struct and f_raw:
        if st.button("Run QC Validation"):
            with st.spinner("Analyzing datasets..."):
                qc_df, error_msg = run_qc_check(f_struct, f_raw)
                
                if error_msg:
                    st.error(error_msg)
                else:
                    st.success("Validation Complete!")
                    
                    # --- Metrics ---
                    avg_acc = qc_df["Accuracy (%)"].mean()
                    perfect_cols = len(qc_df[qc_df["Accuracy (%)"] == 100])
                    total_cols = len(qc_df)
                    
                    m1, m2, m3 = st.columns(3)
                    m1.metric("Overall Average Accuracy", f"{avg_acc:.2f}%")
                    m2.metric("Perfectly Matched Columns", f"{perfect_cols} / {total_cols}")
                    m3.metric("Data Source", "User Uploaded")
                    
                    st.markdown("---")
                    st.subheader("Detailed Column Analysis")
                    
                    # Highlight low accuracy rows
                    def highlight_low_acc(val):
                        color = 'red' if val < 100 else 'green'
                        return f'color: {color}'

                    st.dataframe(
                        qc_df.style.applymap(highlight_low_acc, subset=['Accuracy (%)']),
                        use_container_width=True,
                        height=600
                    )
                    
                    # CSV Download
                    csv = qc_df.to_csv(index=False).encode('utf-8')
                    st.download_button(
                        label="📥 Download QC Report (CSV)",
                        data=csv,
                        file_name='QC_Validation_Report.csv',
                        mime='text/csv',
                    )